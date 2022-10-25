import csv
from openpyxl import load_workbook

PARC_NAME = []

WSUS = []
BITDEFENDER = []
QUALIAC = []
FIDUCIA = []
GIPOM = []
LANSWEEPER = []
SCCM = []
OU = []
UPDATE = []


def init():
    fiducia = load_workbook('../sources/fiducia.xlsx', read_only=True)
    sheet_fiducia = fiducia.worksheets[0]

    gipom = load_workbook('../sources/gipom.xlsx', read_only=True)
    sheet_gipom = gipom.worksheets[0]

    qualiac = load_workbook('../sources/qualiac.xlsx', read_only=True)
    sheet_qualiac = qualiac.worksheets[0]

    lansweeper = load_workbook('../sources/lansweeper.xlsx', read_only=True)
    sheet_lansweeper = lansweeper.worksheets[0]

    wsus = load_workbook('../sources/wsus.xlsx', read_only=True)
    sheet_wsus = wsus.worksheets[0]

    sccm = load_workbook('../sources/sccm.xlsx', read_only=True)
    sheet_sccm = sccm.worksheets[0]

    ou = load_workbook('../sources/ou.xlsx', read_only=True)
    sheet_ou = ou.worksheets[0]

    winupdate = load_workbook('../sources/update.xlsx', read_only=True)
    sheet_winupdate = winupdate.worksheets[2]

    csvfile = open("../sources/bitdefender.csv")
    csvreader = csv.reader(csvfile)

    for row in sheet_lansweeper.rows:
        if row[0].value is not None and row[10].value is not None and row[15].value is not None:
            if row[0].value[0].upper() == "L" or row[0].value[0].upper() == "W":
                LANSWEEPER.append([row[0].value.upper(), row[10].value.upper(), row[15].value])
                PARC_NAME.append(row[0].value.upper())

    for row in sheet_ou.rows:
        if row[0].value is not None and row[11].value is not None:
            OU.append([row[0].value.upper(), row[11].value.upper()])

    for row in csvreader:
        if row[0] is not None and row[5] is not None:
            BITDEFENDER.append([row[0].upper(), row[5]])

    for row in sheet_wsus.rows:
        if row[1].value is not None and row[4].value is not None:
            WSUS.append([row[1].value.upper()[:-6], row[4].value])

    for row in sheet_sccm.rows:
        if row[1].value is not None:
            SCCM.append(row[1].value)

    for row in sheet_fiducia.rows:
        if row[0].value is not None:
            FIDUCIA.append(row[0].value.upper())

    for row in sheet_gipom.rows:
        if row[0].value is not None:
            GIPOM.append(row[0].value.upper())

    for row in sheet_qualiac.rows:
        if row[0].value is not None:
            QUALIAC.append(row[0].value.upper())

    for row in sheet_winupdate.rows:
        list_updates = []
        for el in row:
            list_updates.append(el.value)
        if len(list_updates) == 6:
            if list_updates[3] is not None and list_updates[5] is not None:
                UPDATE.append([list_updates[3].upper()[:-6], list_updates[5].upper()])

    csvfile.close()
    lansweeper.close()
    fiducia.close()
    gipom.close()
    qualiac.close()
    wsus.close()
    sccm.close()
    ou.close()
    winupdate.close()
