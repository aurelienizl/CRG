from linker import *
from datetime import date
from openpyxl import *
from openpyxl.styles import PatternFill

NON_CONFORME = []

today = date.today()
file_name = today.strftime("%b-%d-%Y")


class Computer:
    name = ""
    location = ""
    type = ""
    av = ""
    crit = ""
    wsus = ""
    sccm = ""
    ou = ""
    alive = ""
    os = ""
    update = ""
    bulle = ""
    remediation = ""


PARC = []


def typedef(name):
    name = name[0]
    if name == 'L':
        return "Laptop"
    if name == 'W':
        return "Desktop - VM"
    print("UNABLE TO DETERMINE TYPE OF " + name)
    return "UNDEFINED"


def locationdef(name):
    identity = name[1] + name[2]
    if identity.isnumeric():
        number = identity
        if number == "00":
            return "Wallis"
        if number == "10":
            return "Guadeloupe"
        if number == "20":
            return "Guyane"
        if number == "30":
            return "Martinique"
        if number == "40":
            return "Mayotte"
        if number == "50":
            return "Nouvelle Caledonie"
        if number == "60":
            return "Polynesie"
        if number == "70":
            return "Reunion"
        if number == "80":
            return "SPM"
        if number == "90":
            return "Paris"
        else:
            return "UNDEFINED"
    else:
        print("Unable to find location of computer: " + name)
        return "UNDEFINED"


def avdef(name):
    if name in BITDEFENDER_STATUS:
        return "WARN"
    else:
        return "OK"


def critdef(name):
    if name in PC_CRITIQUE:
        return "CRITIQUE"
    else:
        return "BUREAUTIQUE"


def wsusdef(name):
    if name in WSUS_STATUS:
        return "WARN"

    else:
        return "OK"


def sccmdef(name):
    if name in SCCM_STATUS:
        return "WARN"
    else:
        return "OK"


def oudef(name):
    if name in OU_STATUS:
        return "WARN"
    else:
        return "OK"


def updatedef(name):
    if name in UPDATE_STATUS:
        return "WARN"
    else:
        return "OK"


def bulledef(name):
    if name in BULLE:
        return "POSTE ISOLE"
    else:
        return "NON ISOLE"


def remediationdef(name):
    if name in REMEDIATION:
        return "EN REMEDIATION"
    else:
        return "NON"


def generate_data():
    for el in LANSWEEPER_STATUS:
        pc = Computer()

        pc.name = el[0]
        pc.os = el[1]
        pc.alive = el[2]

        pc.location = locationdef(el[0])
        pc.type = typedef(el[0])
        pc.av = avdef(el[0])
        pc.crit = critdef(el[0])
        pc.wsus = wsusdef(el[0])
        pc.sccm = sccmdef(el[0])
        pc.ou = oudef(el[0])
        pc.update = updatedef(el[0])
        pc.bulle = bulledef(el[0])
        pc.remediation = remediationdef(el[0])

        PARC.append(pc)


def check_issue(pc):
    if pc.os != "WIN 7":
        if pc.remediation == "NON":
            if "WARN" in pc.av:
                return True
            if "WARN" in pc.wsus:
                return True
            if "WARN" in pc.sccm:
                return True
            if "WARN" in pc.ou:
                return True
            if "WARN" in pc.update:
                return True
    return 0


def write_data():
    file = Workbook(write_only=True)

    overview = file.create_sheet()
    overview.title = "OVERVIEW"
    overview.append(
        ["NAME", "TYPE", "LOCATION", "AV", "WSUS", "SCCM", "OU", "UPDATE", "CRITICITE", "OS", "BULLE", "REMEDIATION",
         "ALIVE"])
    for el in PARC:
        overview.append(
            [el.name, el.type, el.location, el.av, el.wsus, el.sccm, el.ou, el.update, el.crit, el.os, el.bulle,
             el.remediation, el.alive])

    issues = file.create_sheet()
    issues.title = "ISSUES"
    issues.append(
        ["NAME", "TYPE", "LOCATION", "AV", "WSUS", "SCCM", "OU", "UPDATE", "CRITICITE", "OS", "BULLE", "REMEDIATION",
         "ALIVE"])
    for el in PARC:
        if check_issue(el):
            NON_CONFORME.append(
                [el.name, el.type, el.location, el.av, el.wsus, el.sccm, el.ou, el.update, el.crit, el.os, el.bulle,
                 el.remediation, el.alive])
            issues.append(
                [el.name, el.type, el.location, el.av, el.wsus, el.sccm, el.ou, el.update, el.crit, el.os, el.bulle,
                 el.remediation, el.alive])

    file.save(filename="../conformity-report/" + file_name + ".xlsx")
    file.close()


def color():
    data = load_workbook("../conformity-report/" + file_name + ".xlsx")
    for sheet_data in data.worksheets:
        for row in sheet_data.rows:
            for el in row:
                if el.value == "WARN":
                    el.fill = PatternFill('solid', fgColor='FFFF0000')
                elif el.value == "CRITIQUE":
                    el.fill = PatternFill('solid', fgColor='FF7B00')
                elif el.value == "WIN 10" or el.value == "BUREAUTIQUE" or el.value == "NON":
                    el.fill = PatternFill('solid', fgColor='ADD8E6')
                elif el.value == "WIN 7":
                    el.fill = PatternFill('solid', fgColor='FFFFFF00')
                elif el.value == "OK":
                    el.fill = PatternFill('solid', fgColor='FF00FF00')
                elif el.value == "NON ISOLE":
                    el.fill = PatternFill('solid', fgColor='ADD8E6')
                elif el.value == "POSTE ISOLE" or el.value == "EN REMEDIATION":
                    el.fill = PatternFill('solid', fgColor='FF7B00')

    data.save(filename="../conformity-report/" + file_name + ".xlsx")
    data.close()


def class_default():
    data = load_workbook("../conformity-report/" + file_name + ".xlsx")
    bureautique_issues = data.create_sheet()
    bureautique_issues.title = "BUREAUTIQUE"
    critique_issues = data.create_sheet()
    critique_issues.title = "CRITIQUE"
    bureautique_issues.append(
        ["NAME", "TYPE", "LOCATION", "AV", "WSUS", "SCCM", "OU", "UPDATE", "CRITICITE", "OS", "BULLE", "REMEDIATION",
         "ALIVE"])
    critique_issues.append(
        ["NAME", "TYPE", "LOCATION", "AV", "WSUS", "SCCM", "OU", "UPDATE", "CRITICITE", "OS", "BULLE", "REMEDIATION",
         "ALIVE"])
    for el in NON_CONFORME:
        if "BU" in el[8]:
            bureautique_issues.append(el)
        else:
            critique_issues.append(el)

    data.save(filename="../conformity-report/" + file_name + ".xlsx")
    data.close()


def generate_rapport():
    init()
    status()
    generate_data()
    write_data()
    class_default()
    color()


generate_rapport()
